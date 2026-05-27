#!/usr/bin/env python3
"""Convert judge-validation CSVs into Excel-friendly XLSX with text wrapping.

Usage:
    python3 scripts/csv_to_xlsx.py judge_validation_computational.csv judge_validation_medical.csv
"""

import csv
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WIDE_COLS = {"question", "correct_answer", "model_response"}
NARROW_COLS = {"question_id", "challenge_id", "condition", "model"}
HIDDEN_FROM_RATER = {"judge_factual_accuracy", "judge_agreement"}
RATER_COLS = {"human_factual_accuracy", "human_agreement"}


def convert(csv_path: str) -> str:
    out_path = csv_path.rsplit(".", 1)[0] + ".xlsx"
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        print(f"Empty CSV: {csv_path}")
        return out_path

    wb = Workbook()
    ws = wb.active
    ws.title = "annotation"

    headers = list(rows[0].keys())
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    rater_fill = PatternFill("solid", fgColor="FFF2CC")
    judge_fill = PatternFill("solid", fgColor="F4CCCC")

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    wrap_align = Alignment(wrap_text=True, vertical="top")
    plain_align = Alignment(vertical="top")
    for col_idx, h in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        if h in WIDE_COLS:
            ws.column_dimensions[letter].width = 60
            for row_idx in range(2, len(rows) + 2):
                ws.cell(row=row_idx, column=col_idx).alignment = wrap_align
        elif h in NARROW_COLS:
            ws.column_dimensions[letter].width = 16
            for row_idx in range(2, len(rows) + 2):
                ws.cell(row=row_idx, column=col_idx).alignment = plain_align
        elif h in RATER_COLS:
            ws.column_dimensions[letter].width = 22
            for row_idx in range(2, len(rows) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = rater_fill
                cell.alignment = plain_align
        elif h in HIDDEN_FROM_RATER:
            ws.column_dimensions[letter].width = 22
            ws.column_dimensions[letter].hidden = True
            for row_idx in range(2, len(rows) + 2):
                ws.cell(row=row_idx, column=col_idx).fill = judge_fill
        else:
            ws.column_dimensions[letter].width = 18

    ws.row_dimensions[1].height = 30
    response_col_idx = headers.index("model_response") + 1 if "model_response" in headers else None
    for row_idx, r in enumerate(rows, start=2):
        if response_col_idx is None:
            ws.row_dimensions[row_idx].height = 120
            continue
        resp = r.get("model_response", "") or ""
        approx_lines = max(1, len(resp) // 60 + resp.count("\n"))
        height = min(600, max(120, approx_lines * 15))
        ws.row_dimensions[row_idx].height = height

    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Wrote {out_path} ({len(rows)} rows)")
    return out_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/csv_to_xlsx.py <csv> [<csv> ...]")
        sys.exit(1)
    for path in sys.argv[1:]:
        convert(path)


if __name__ == "__main__":
    main()
